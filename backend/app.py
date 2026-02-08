from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import pyodbc
import pandas as pd
import json
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for React frontend

# Load Configuration
def load_config():
    config_path = os.path.join(os.path.dirname(__file__), 'db_config.json')
    try:
        with open(config_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {"environments": []}

CONFIG = load_config()

import uuid
from comparison_engine import run_hybrid_comparison
from storage_manager import save_df, load_df, clear_cache

# Initialize Cache (Optional: clear old cache on restart)
# clear_cache() 

@app.route('/api/config', methods=['GET'])
def get_config():
    """Returns the environment hierarchy for the frontend dropdowns."""
    return jsonify(CONFIG)

def get_connection_string(server, database):
    return (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        "Trusted_Connection=yes;"
    )

@app.route('/api/preview_sql', methods=['POST'])
def preview_sql():
    """
    Executes a SQL query and returns columns + top 5 rows.
    """
    data = request.json
    server = data.get('server')
    database = data.get('database')
    query = data.get('query')

    if not server or not database or not query:
        return jsonify({"error": "Missing parameters"}), 400
        
    # Basic Safety Check (Prevent modifications)
    forbidden_keywords = ['DROP', 'DELETE', 'UPDATE', 'INSERT', 'TRUNCATE', 'ALTER', 'GRANT', 'REVOKE', 'EXEC', 'CREATE', 'MERGE']
    if any(keyword in query.upper() for keyword in forbidden_keywords):
        return jsonify({"error": "Security Alert: Only SELECT queries are permitted in this environment."}), 403

    conn_str = get_connection_string(server, database)

    try:
        conn = pyodbc.connect(conn_str, timeout=10)
        # using pandas read_sql to easy fetch
        # Use simple query wrapper to fetch limited rows for preview if user didn't specify TOP
        preview_query = query
        
        # This is a basic preview, we fetch everything for now to get types, 
        # but for 1M rows in production we would want to wrap this in a top 10 logic for the preview endpoint specificially.
        # For this tool, let's just fetch top 10 for the preview.
        # Note: Modifying user query is risky, so we just run it and head(5) in pandas if it's not too huge.
        # Better approach for preview:
        if "TOP" not in query.upper() and "LIMIT" not in query.upper():
             # Basic injection of TOP 5 for preview purposes if MSSQL
             # A robust parser is complex, so we will read with chunksize
            pass

        # Read only small chunk for preview
        df = pd.read_sql(query, conn) #, chunksize=5) 
        # In a real scenario with 1M rows, read_sql without chunking is slow for a "Preview". 
        # But for 'preview', the user usually writes 'SELECT TOP 10 * ...'
        # Let's assume user writes the query. If they write 'Select *', it might be slow.
        
        preview_df = df.head(5)
        columns = list(preview_df.columns)
        rows = preview_df.to_dict(orient='records')
        
        # Include last row for truncated preview display in console
        last_row = None
        if len(df) > 5:
            last_row = df.tail(1).fillna("").to_dict(orient='records')[0]
        
        # Cache the Query logic? No, we will re-run full query on 'Run Comparison'
        
        conn.close()
        return jsonify({
            "status": "success",
            "columns": columns,
            "preview_data": rows,
            "row_count_estimate": len(df),
            "last_row": last_row
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/upload_file', methods=['POST'])
def upload_file():
    """
    Uploads a file, saves it temp, and returns columns + preview.
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    unique_id = str(uuid.uuid4())
    filename = file.filename
    
    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file)
        else:
             return jsonify({"error": "Invalid file format. Only CSV or Excel allowed."}), 400
        
        # Save to Disk Cache FIRST (Parquet handles NaN natively)
        save_df(df, 'uploads', unique_id)
        
        # Then build preview (convert NaN to None only for JSON response)
        preview_df = df.head(5).fillna("")
        columns = list(preview_df.columns)
        rows = preview_df.to_dict(orient='records')
        
        return jsonify({
            "status": "success",
            "file_id": unique_id,
            "columns": columns,
            "preview_data": rows,
            "total_rows": len(df)
        })
        
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/connect', methods=['POST'])
def test_connection():
    """
    Tests connection to SQL Server using Windows Authentication.
    Expects JSON: { "server": "1.2.3.4", "database": "MyDB" }
    """
    data = request.json
    server = data.get('server')
    database = data.get('database')
    
    if not server or not database:
        return jsonify({"error": "Missing server or database parameters"}), 400

    # Connection String for Windows Auth (Trusted Connection)
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        "Trusted_Connection=yes;"
    )

    try:
        # Check installed drivers for debugging
        drivers = [x for x in pyodbc.drivers() if 'SQL Server' in x]
        
        # Test connection with a short timeout
        conn = pyodbc.connect(conn_str, timeout=5)
        conn.close()
        return jsonify({
            "status": "success", 
            "message": f"Successfully connected to {database} on {server}",
            "info": "Authenticated via Windows Trusted Connection"
        })
    
    except pyodbc.Error as ex:
        # Extract SQL State if available
        sqlstate = ex.args[0] if ex.args else 'UNKNOWN'
        error_msg = str(ex)
        
        # Detailed Error Diagnosis
        detailed_reason = "An unexpected error occurred during connection."
        
        if '08001' in str(ex) or '08001' in str(sqlstate):
            detailed_reason = (
                "Unreachable Server (Network/Instance Error). "
                "Possible causes: "
                "1. The server IP/Hostname is incorrect. "
                "2. SQL Server is not running on the target. "
                "3. Firewall is blocking port 1433. "
                "4. SQL Browser service is down."
            )
        elif '28000' in str(ex) or '28000' in str(sqlstate):
            detailed_reason = (
                "Authentication Failed. "
                "You are using Windows Authentication (Trusted_Connection=yes). "
                "Your current Windows User does not have permission to access this database. "
                "Please Request Access or check if you are logged in as the correct user."
            )
        elif 'HYT00' in str(ex) or 'HYT00' in str(sqlstate):
            detailed_reason = (
                "Connection Timeout. "
                "The server exists but took too long to respond. "
                "The network might be slow or the server is under heavy load."
            )
            
        return jsonify({
            "status": "error", 
            "code": sqlstate, 
            "message": detailed_reason,
            "raw_error": error_msg,
            "available_drivers": drivers
        }), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/run_comparison', methods=['POST'])
def run_comparison():
    """
    Orchestrates the Comparison:
    1. Fetches FULL SQL data.
    2. Retrieves Uploaded File data from Cache.
    3. Runs Hybrid Comparison Engine.
    4. Caches Result.
    """
    data = request.json
    file_id = data.get('file_id')
    server = data.get('server')
    database = data.get('database')
    query = data.get('query')
    keys = data.get('keys', []) # List of SQL-side column names used as keys
    column_mapping = data.get('column_mapping', []) # [{sql: "name", file: "student name"}, ...]

    if not file_id or not server or not database or not query:
        return jsonify({"error": "Missing required parameters"}), 400

    # 1. Retrieve File Data
    df_file = load_df('uploads', file_id)
    if df_file is None:
        return jsonify({"error": "File session expired or invalid. Please re-upload."}), 404

    try:
        # 2. Fetch Full SQL Data
        conn_str = get_connection_string(server, database)
        conn = pyodbc.connect(conn_str)
        df_sql = pd.read_sql(query, conn)
        conn.close()

        # 3. Apply Column Mapping
        # column_mapping tells us which file columns correspond to which SQL columns.
        # We rename file columns to match SQL column names, and keep only mapped columns.
        if column_mapping and len(column_mapping) > 0:
            # Build rename dict: file_col -> sql_col
            rename_map = {m['file']: m['sql'] for m in column_mapping}
            mapped_sql_cols = [m['sql'] for m in column_mapping]
            mapped_file_cols = [m['file'] for m in column_mapping]

            # Filter SQL df to only mapped columns
            df_sql = df_sql[[c for c in mapped_sql_cols if c in df_sql.columns]]

            # Filter File df to only mapped columns, then rename to SQL names
            df_file = df_file[[c for c in mapped_file_cols if c in df_file.columns]]
            df_file = df_file.rename(columns=rename_map)

        # 4. Run Logic
        result_df, summary = run_hybrid_comparison(df_sql, df_file, keys)
        
        # 4. Cache Result
        result_id = str(uuid.uuid4())
        save_df(result_df, 'results', result_id)
        
        # 5. Return Summary + First Page
        # Replace NaN with null for valid JSON
        preview_page = result_df.head(50).fillna("").to_dict(orient='records')
        
        return jsonify({
            "status": "success",
            "result_id": result_id,
            "summary": summary,
            "preview_rows": preview_page,
            "columns": list(result_df.columns)
        })

    except Exception as e:
        # cleanup if error?
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/results_page', methods=['GET'])
def get_results_page():
    """ 
    Pagination for the Data Grid
    Query Params: result_id, page (1-based), size (default 100)
    """
    result_id = request.args.get('result_id')
    page = int(request.args.get('page', 1))
    size = int(request.args.get('size', 100))
    
    df = load_df('results', result_id)
    if df is None:
        return jsonify({"error": "Result cache expired. Run comparison again."}), 404
        
    # Calculate slice
    start = (page - 1) * size
    end = start + size
    
    # Slice
    # Handle end of list
    if start >= len(df):
        return jsonify({"data": [], "page": page, "has_more": False})
        
    sliced_df = df.iloc[start:end]
    data = sliced_df.fillna("").to_dict(orient='records')
    
    return jsonify({
        "data": data,
        "page": page,
        "total_pages": (len(df) // size) + 1,
        "has_more": end < len(df)
    })

@app.route('/api/export_excel', methods=['GET'])
def export_excel():
    """Generate styled Excel file with color-coded reconciliation results."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    result_id = request.args.get('result_id')
    if not result_id:
        return jsonify({"error": "Missing result_id"}), 400

    df = load_df('results', result_id)
    if df is None:
        return jsonify({"error": "Result cache expired. Run comparison again."}), 404

    df = df.fillna('')

    # ── Color definitions (match web UI) ──
    header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    pre_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')       # yellow-100 (pre/SQL)
    post_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')      # green-100 (post/File)
    mismatch_cell_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')  # red-200 (changed cell)
    mismatch_font = Font(bold=True, color='7F1D1D')
    sql_only_fill = PatternFill(start_color='FDE68A', end_color='FDE68A', fill_type='solid')   # amber-200
    file_only_fill = PatternFill(start_color='FECDD3', end_color='FECDD3', fill_type='solid')  # rose-200
    section_title_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    section_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    display_cols = [c for c in df.columns if c != '_mismatch_cols']

    # ── Split data by category ──
    mismatched = df[df['status'] == 'Mismatch']
    missing_df = df[df['status'] == 'Only in SQL']
    extra_df = df[df['status'] == 'Only in File']

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation Results"

    def write_section(ws, section_df, title, start_row):
        """Write a section with title, headers, and colored data rows."""
        # Section title row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(display_cols))
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_title_fill
        start_row += 1

        # Column headers
        for ci, col_name in enumerate(display_cols, 1):
            cell = ws.cell(row=start_row, column=ci, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        start_row += 1

        # Data rows
        for _, row in section_df.iterrows():
            mismatch_cols = [c.strip() for c in str(row.get('_mismatch_cols', '')).split(',') if c.strip()]
            status = str(row.get('status', ''))
            pre_post = str(row.get('pre/post', ''))

            # Determine base row fill color
            if status == 'Mismatch':
                base_fill = pre_fill if pre_post == 'pre' else post_fill
            elif status == 'Only in SQL':
                base_fill = sql_only_fill
            elif status == 'Only in File':
                base_fill = file_only_fill
            else:
                base_fill = None

            for ci, col_name in enumerate(display_cols, 1):
                val = row.get(col_name, '')
                cell = ws.cell(row=start_row, column=ci, value=str(val))
                cell.border = thin_border

                if base_fill:
                    cell.fill = base_fill

                # Highlight mismatched cells with red
                if col_name in mismatch_cols and status == 'Mismatch':
                    cell.fill = mismatch_cell_fill
                    cell.font = mismatch_font

            start_row += 1

        return start_row + 1  # gap between sections

    current_row = 1

    if len(mismatched) > 0:
        current_row = write_section(ws, mismatched, f"Mismatched Rows ({len(mismatched) // 2})", current_row)
    if len(missing_df) > 0:
        current_row = write_section(ws, missing_df, f"Missing from File / SQL Only ({len(missing_df)})", current_row)
    if len(extra_df) > 0:
        current_row = write_section(ws, extra_df, f"Extra in File / Not in SQL ({len(extra_df)})", current_row)

    if len(df) == 0:
        ws.cell(row=1, column=1, value="No discrepancies found - data matches perfectly!")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color='16A34A')

    # Auto-width columns
    for ci, col_name in enumerate(display_cols, 1):
        col_letter = get_column_letter(ci)
        max_len = len(str(col_name))
        for i, (_, row) in enumerate(df.iterrows()):
            if i >= 100:  # sample first 100 rows
                break
            val_len = len(str(row.get(col_name, '')))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Save to memory and return
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'reconciliation_{result_id[:8]}.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)
